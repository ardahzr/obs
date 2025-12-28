from django.shortcuts import render
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db.models import Avg, Sum, F, Q
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from rest_framework.authtoken.models import Token
from django.utils import timezone
from django.http import HttpResponse
from collections import Counter
import io

from .models import (
    Course, ProgramOutcome, LearningOutcome, 
    LoToPoMapping, Student, Assessment, AssessmentToLoMapping, Grade, UserProfile, Notification, Enrollment
)
from .serializers import (
    CourseSerializer, CourseDetailSerializer, ProgramOutcomeSerializer,
    LearningOutcomeSerializer, LoToPoMappingSerializer, StudentSerializer,
    AssessmentSerializer, AssessmentToLoMappingSerializer, GradeSerializer,
    LoginSerializer, RegisterSerializer, UserSerializer, NotificationSerializer
)
from .chat_utils import chat_with_gemini
from .email_utils import send_notification_email


# Create your views here.

@api_view(['GET'])
def test_api(request):
    return Response({"message": "Hello from Django API!"})


# ============ AUTH VIEWS ============

import requests as http_requests
from django.conf import settings as django_settings

def verify_recaptcha(token):
    """Verify reCAPTCHA token with Google"""
    if not token:
        return False
    
    secret_key = django_settings.RECAPTCHA_SECRET_KEY
    if not secret_key:
        # If no secret key configured, skip verification (development mode)
        return True
    
    try:
        response = http_requests.post(
            'https://www.google.com/recaptcha/api/siteverify',
            data={
                'secret': secret_key,
                'response': token
            },
            timeout=5
        )
        result = response.json()
        return result.get('success', False)
    except Exception as e:
        print(f"reCAPTCHA verification error: {e}")
        return False


@api_view(['POST'])
def login_view(request):
    """User login"""
    # Verify reCAPTCHA
    recaptcha_token = request.data.get('recaptcha_token')
    if not verify_recaptcha(recaptcha_token):
        return Response({
            'success': False,
            'message': 'reCAPTCHA doğrulaması başarısız. Lütfen tekrar deneyin.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        
        # Get profile info
        try:
            profile = user.profile
            user_type = profile.user_type
        except UserProfile.DoesNotExist:
            user_type = 'instructor'
        
        return Response({
            'success': True,
            'message': 'Login successful',
            'token': token.key,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'user_type': user_type
            }
        })
    return Response({
        'success': False,
        'message': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def logout_view(request):
    """User logout"""
    try:
        # Delete token
        if request.auth:
            request.auth.delete()
    except:
        pass
    
    return Response({
        'success': True,
        'message': 'Logout successful'
    })


@api_view(['POST'])
def register_view(request):
    """New user registration (admin only)"""
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        return Response({
            'success': True,
            'message': 'User created successfully',
            'user': UserSerializer(user).data
        }, status=status.HTTP_201_CREATED)
    return Response({
        'success': False,
        'message': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
def me_view(request):
    """Mevcut kullanıcı bilgisi"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Token '):
        return Response({
            'success': False,
            'message': 'Token gerekli'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    token_key = auth_header.split(' ')[1]
    try:
        token = Token.objects.get(key=token_key)
        user = token.user
        
        try:
            profile = user.profile
            user_type = profile.user_type
        except UserProfile.DoesNotExist:
            user_type = 'instructor'
        
        return Response({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'user_type': user_type
            }
        })
    except Token.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Geçersiz token'
        }, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
def create_instructor(request):
    """
    Admin tarafından instructor hesabı oluştur.
    Rastgele şifre oluşturur ve email ile gönderir.
    """
    import secrets
    import string
    
    user = get_user_from_token(request)
    if not user:
        return Response({'success': False, 'message': 'Yetkilendirme gerekli'}, status=401)
    
    # Admin kontrolü
    try:
        if user.profile.user_type != 'admin':
            return Response({'success': False, 'message': 'Admin yetkisi gerekli'}, status=403)
    except UserProfile.DoesNotExist:
        return Response({'success': False, 'message': 'Admin yetkisi gerekli'}, status=403)
    
    # Gerekli alanları al
    username = request.data.get('username', '').strip()
    email = request.data.get('email', '').strip()
    first_name = request.data.get('first_name', '').strip()
    last_name = request.data.get('last_name', '').strip()
    
    if not username or not email:
        return Response({
            'success': False,
            'message': 'Kullanıcı adı ve email zorunludur'
        }, status=400)
    
    # Kullanıcı adı veya email zaten var mı kontrol et
    if User.objects.filter(username=username).exists():
        return Response({
            'success': False,
            'message': 'Bu kullanıcı adı zaten kullanılıyor'
        }, status=400)
    
    if User.objects.filter(email=email).exists():
        return Response({
            'success': False,
            'message': 'Bu email adresi zaten kullanılıyor'
        }, status=400)
    
    # Rastgele şifre oluştur (12 karakter)
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    password = ''.join(secrets.choice(alphabet) for _ in range(12))
    
    # Kullanıcı oluştur
    new_user = User.objects.create_user(
        username=username,
        email=email,
        password=password,
        first_name=first_name,
        last_name=last_name
    )
    
    # UserProfile oluştur (instructor olarak)
    UserProfile.objects.create(user=new_user, user_type='instructor')
    
    # Email gönder
    from django.core.mail import send_mail
    from django.conf import settings as django_settings
    
    try:
        html_message = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #667eea, #764ba2); color: white; padding: 20px; border-radius: 10px 10px 0 0; text-align: center; }}
                .content {{ background: #f9fafb; padding: 25px; border: 1px solid #e5e7eb; }}
                .credentials {{ background: #1a1a2e; color: #fff; padding: 20px; border-radius: 10px; margin: 20px 0; }}
                .credentials p {{ margin: 10px 0; }}
                .label {{ color: #a0aec0; font-size: 12px; }}
                .value {{ font-size: 18px; font-weight: bold; color: #667eea; }}
                .footer {{ background: #f3f4f6; padding: 15px; text-align: center; font-size: 12px; color: #6b7280; border-radius: 0 0 10px 10px; }}
                .warning {{ background: #fef3c7; border: 1px solid #f59e0b; padding: 12px; border-radius: 8px; margin-top: 15px; color: #92400e; font-size: 13px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1 style="margin: 0;">🎓 PO Manager</h1>
                    <p style="margin: 5px 0 0 0; opacity: 0.9;">Program Outcome Management System</p>
                </div>
                <div class="content">
                    <h2>Hoş Geldiniz, {first_name or username}!</h2>
                    <p>PO Manager sistemine instructor olarak kaydınız oluşturulmuştur. Aşağıdaki bilgilerle giriş yapabilirsiniz:</p>
                    
                    <div class="credentials">
                        <p><span class="label">KULLANICI ADI</span><br><span class="value">{username}</span></p>
                        <p><span class="label">ŞİFRE</span><br><span class="value">{password}</span></p>
                    </div>
                    
                    <div class="warning">
                        ⚠️ Güvenliğiniz için ilk girişten sonra şifrenizi değiştirmenizi öneririz.
                    </div>
                </div>
                <div class="footer">
                    <p>Bu email PO Manager sistemi tarafından otomatik olarak gönderilmiştir.</p>
                    <p>© 2025 PO Manager - Program Outcome Management System</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        send_mail(
            subject='🎓 PO Manager - Hesap Bilgileriniz',
            message=f'Kullanıcı adı: {username}\nŞifre: {password}',
            from_email=django_settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            html_message=html_message,
            fail_silently=False,
        )
        email_sent = True
    except Exception as e:
        print(f"Email gönderme hatası: {e}")
        email_sent = False
    
    return Response({
        'success': True,
        'message': 'Instructor hesabı başarıyla oluşturuldu',
        'email_sent': email_sent,
        'user': {
            'id': new_user.id,
            'username': new_user.username,
            'email': new_user.email,
            'first_name': new_user.first_name,
            'last_name': new_user.last_name,
            'user_type': 'instructor'
        }
    }, status=201)


class CourseViewSet(viewsets.ModelViewSet):
    """Ders CRUD işlemleri"""
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    
    @action(detail=True, methods=['get'])
    def detail(self, request, pk=None):
        """Detaylı ders bilgisi (LO'lar ve assessments dahil)"""
        course = self.get_object()
        serializer = CourseDetailSerializer(course)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def learning_outcomes(self, request, pk=None):
        """Dersin learning outcomes listesi"""
        course = self.get_object()
        los = course.learning_outcomes.all()
        serializer = LearningOutcomeSerializer(los, many=True)
        return Response(serializer.data)


class ProgramOutcomeViewSet(viewsets.ModelViewSet):
    """Program Outcome CRUD işlemleri"""
    queryset = ProgramOutcome.objects.all()
    serializer_class = ProgramOutcomeSerializer


class LearningOutcomeViewSet(viewsets.ModelViewSet):
    """Learning Outcome CRUD işlemleri"""
    queryset = LearningOutcome.objects.all()
    serializer_class = LearningOutcomeSerializer
    
    @action(detail=True, methods=['get', 'post'])
    def mappings(self, request, pk=None):
        """LO'nun PO mappings'leri"""
        lo = self.get_object()
        
        if request.method == 'GET':
            mappings = lo.po_mappings.all()
            serializer = LoToPoMappingSerializer(mappings, many=True)
            return Response(serializer.data)
        
        elif request.method == 'POST':
            # Yeni mapping oluştur
            data = request.data.copy()
            data['learning_outcome'] = lo.id
            serializer = LoToPoMappingSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoToPoMappingViewSet(viewsets.ModelViewSet):
    """LO to PO Mapping CRUD işlemleri"""
    queryset = LoToPoMapping.objects.all()
    serializer_class = LoToPoMappingSerializer


class AssessmentToLoMappingViewSet(viewsets.ModelViewSet):
    """Assessment to LO Mapping CRUD işlemleri"""
    queryset = AssessmentToLoMapping.objects.all()
    serializer_class = AssessmentToLoMappingSerializer


class StudentViewSet(viewsets.ModelViewSet):
    """Öğrenci CRUD işlemleri"""
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    
    def get_queryset(self):
        """Ders bazında filtreleme desteği"""
        queryset = Student.objects.all()
        course_id = self.request.query_params.get('course')
        if course_id:
            queryset = queryset.filter(enrollments__course_id=course_id)
        return queryset.distinct()
    
    @action(detail=True, methods=['get'])
    def grades(self, request, pk=None):
        """Öğrencinin notları"""
        student = self.get_object()
        grades = student.grades.all()
        serializer = GradeSerializer(grades, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def po_scores(self, request, pk=None):
        """
        Öğrencinin PO skorlarını hesapla
        Formül: (Assessment_Score * Assessment_Weight * LO_Weight) / Total_Weight
        """
        student = self.get_object()
        
        # Tüm PO'ları al
        pos = ProgramOutcome.objects.all()
        results = []
        
        for po in pos:
            # Bu PO'ya bağlı tüm LO mappinglerini bul
            lo_mappings = LoToPoMapping.objects.filter(program_outcome=po)
            
            total_weighted_score = 0
            total_weight_sum = 0
            
            for lo_map in lo_mappings:
                lo = lo_map.learning_outcome
                lo_po_weight = lo_map.contribution_weight  # LO -> PO ağırlığı (0-1 arası)
                
                # Bu LO'ya bağlı assessment mappinglerini bul
                assess_mappings = AssessmentToLoMapping.objects.filter(learning_outcome=lo)
                
                for assess_map in assess_mappings:
                    assessment = assess_map.assessment
                    assess_lo_weight = assess_map.contribution_weight  # Assessment -> LO ağırlığı (0-1 arası)
                    
                    # Öğrencinin bu assessment'taki notunu bul
                    try:
                        grade = Grade.objects.get(assessment=assessment, student=student)
                        # Notu 100 üzerinden al
                        score = grade.percentage
                        
                        # Katkıyı hesapla: Not * (Assess->LO) * (LO->PO)
                        contribution = score * assess_lo_weight * lo_po_weight
                        
                        # Toplam ağırlık paydası için: (Assess->LO) * (LO->PO)
                        weight_factor = assess_lo_weight * lo_po_weight
                        
                        total_weighted_score += contribution
                        total_weight_sum += weight_factor
                        
                    except Grade.DoesNotExist:
                        continue
            
            # Normalize et: Toplam Puan / Toplam Ağırlık
            if total_weight_sum > 0:
                normalized_score = total_weighted_score / total_weight_sum
            else:
                normalized_score = 0
            
            results.append({
                'po_code': po.code,
                'po_description': po.description,
                'score': round(normalized_score, 2)
            })
        
        return Response({
            'student': student.student_no,
            'po_scores': results
        })


class AssessmentViewSet(viewsets.ModelViewSet):
    """Değerlendirme CRUD işlemleri"""
    queryset = Assessment.objects.all()
    serializer_class = AssessmentSerializer


class GradeViewSet(viewsets.ModelViewSet):
    """Not CRUD işlemleri"""
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer


@api_view(['POST'])
def chat_view(request):
    """Chatbot endpoint"""
    user_message = request.data.get('message')
    if not user_message:
        return Response({"error": "Message is required"}, status=status.HTTP_400_BAD_REQUEST)
    
    # API Key provided by user
    api_key = "AIzaSyAj7cQxoREu3UEv-JU5jllTLYk6U9E6pM8"
    
    try:
        response_text = chat_with_gemini(user_message, api_key)
        return Response({"response": response_text})
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============ NOTIFICATION & APPROVAL VIEWS ============

def get_user_from_token(request):
    """Token'dan kullanıcıyı al"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Token '):
        return None
    token_key = auth_header.split(' ')[1]
    try:
        token = Token.objects.get(key=token_key)
        return token.user
    except Token.DoesNotExist:
        return None


@api_view(['GET'])
def get_notifications(request):
    """Get user notifications"""
    user = get_user_from_token(request)
    if not user:
        return Response({'success': False, 'message': 'Authorization required'}, status=401)
    
    notifications = Notification.objects.filter(recipient=user).order_by('-created_at')
    serializer = NotificationSerializer(notifications, many=True)
    
    unread_count = notifications.filter(is_read=False).count()
    
    return Response({
        'success': True,
        'notifications': serializer.data,
        'unread_count': unread_count
    })


@api_view(['POST'])
def mark_notification_read(request, notification_id):
    """Mark notification as read"""
    user = get_user_from_token(request)
    if not user:
        return Response({'success': False, 'message': 'Authorization required'}, status=401)
    
    try:
        notification = Notification.objects.get(id=notification_id, recipient=user)
        notification.is_read = True
        notification.save()
        return Response({'success': True, 'message': 'Notification marked as read'})
    except Notification.DoesNotExist:
        return Response({'success': False, 'message': 'Notification not found'}, status=404)


@api_view(['POST'])
def mark_all_notifications_read(request):
    """Mark all notifications as read"""
    user = get_user_from_token(request)
    if not user:
        return Response({'success': False, 'message': 'Authorization required'}, status=401)
    
    Notification.objects.filter(recipient=user, is_read=False).update(is_read=True)
    return Response({'success': True, 'message': 'All notifications marked as read'})


@api_view(['POST'])
def submit_course_for_approval(request, course_id):
    """Submit course for approval (Instructor)"""
    user = get_user_from_token(request)
    if not user:
        return Response({'success': False, 'message': 'Authorization required'}, status=401)
    
    try:
        course = Course.objects.get(id=course_id)
        
        # Check course ownership - allow if instructor not assigned, user is instructor, or user is admin
        is_admin = hasattr(user, 'profile') and user.profile.user_type == 'admin'
        is_owner = course.instructor == user or course.instructor is None
        
        if not is_owner and not is_admin:
            return Response({'success': False, 'message': 'You do not have permission to submit this course'}, status=403)
        
        # Check if already pending
        if course.approval_status == 'pending':
            return Response({'success': False, 'message': 'This course is already pending approval'}, status=400)
        
        # If instructor not assigned, assign the submitting user
        if course.instructor is None:
            course.instructor = user
        
        # Submit course for approval
        course.approval_status = 'pending'
        course.submitted_at = timezone.now()
        course.save()
        
        # Send notification to all admin users
        admins = User.objects.filter(profile__user_type='admin')
        for admin in admins:
            notification = Notification.objects.create(
                recipient=admin,
                sender=user,
                course=course,
                notification_type='approval_request',
                message=f'{user.get_full_name() or user.username} submitted "{course.code} - {course.name}" for your approval.'
            )
            # Send email notification
            send_notification_email(notification)
        
        return Response({
            'success': True, 
            'message': 'Course submitted for approval',
            'approval_status': 'pending'
        })
    except Course.DoesNotExist:
        return Response({'success': False, 'message': 'Course not found'}, status=404)


@api_view(['GET'])
def get_pending_approvals(request):
    """Get pending approvals (Admin)"""
    user = get_user_from_token(request)
    if not user:
        return Response({'success': False, 'message': 'Authorization required'}, status=401)
    
    # Admin check
    try:
        if user.profile.user_type != 'admin':
            return Response({'success': False, 'message': 'Admin permission required'}, status=403)
    except UserProfile.DoesNotExist:
        return Response({'success': False, 'message': 'Admin permission required'}, status=403)
    
    pending_courses = Course.objects.filter(approval_status='pending').select_related('instructor')
    serializer = CourseSerializer(pending_courses, many=True)
    
    return Response({
        'success': True,
        'pending_courses': serializer.data
    })


@api_view(['POST'])
def approve_course(request, course_id):
    """Dersi onayla (Admin)"""
    user = get_user_from_token(request)
    if not user:
        return Response({'success': False, 'message': 'Yetkilendirme gerekli'}, status=401)
    
    # Admin kontrolü
    try:
        if user.profile.user_type != 'admin':
            return Response({'success': False, 'message': 'Admin permission required'}, status=403)
    except UserProfile.DoesNotExist:
        return Response({'success': False, 'message': 'Admin permission required'}, status=403)
    
    message = request.data.get('message', 'Your course has been approved.')
    
    try:
        course = Course.objects.get(id=course_id)
        
        if course.approval_status != 'pending':
            return Response({'success': False, 'message': 'This course is not pending approval'}, status=400)
        
        # Approve the course
        course.approval_status = 'approved'
        course.reviewed_at = timezone.now()
        course.reviewed_by = user
        course.rejection_reason = None
        course.save()
        
        # Send notification to course owner
        if course.instructor:
            notification = Notification.objects.create(
                recipient=course.instructor,
                sender=user,
                course=course,
                notification_type='approved',
                message=f'"{course.code} - {course.name}" has been approved. {message}'
            )
            # Send email notification
            send_notification_email(notification)
        
        return Response({
            'success': True,
            'message': 'Course approved',
            'approval_status': 'approved'
        })
    except Course.DoesNotExist:
        return Response({'success': False, 'message': 'Ders bulunamadı'}, status=404)


@api_view(['POST'])
def reject_course(request, course_id):
    """Reject course (Admin)"""
    user = get_user_from_token(request)
    if not user:
        return Response({'success': False, 'message': 'Authorization required'}, status=401)
    
    # Admin check
    try:
        if user.profile.user_type != 'admin':
            return Response({'success': False, 'message': 'Admin permission required'}, status=403)
    except UserProfile.DoesNotExist:
        return Response({'success': False, 'message': 'Admin permission required'}, status=403)
    
    reason = request.data.get('reason', '')
    if not reason:
        return Response({'success': False, 'message': 'Rejection reason is required'}, status=400)
    
    try:
        course = Course.objects.get(id=course_id)
        
        if course.approval_status != 'pending':
            return Response({'success': False, 'message': 'This course is not pending approval'}, status=400)
        
        # Reject the course
        course.approval_status = 'rejected'
        course.reviewed_at = timezone.now()
        course.reviewed_by = user
        course.rejection_reason = reason
        course.save()
        
        # Send notification to course owner
        if course.instructor:
            notification = Notification.objects.create(
                recipient=course.instructor,
                sender=user,
                course=course,
                notification_type='rejected',
                message=f'"{course.code} - {course.name}" has been rejected. Reason: {reason}'
            )
            # Send email notification
            send_notification_email(notification)
        
        return Response({
            'success': True,
            'message': 'Course rejected',
            'approval_status': 'rejected'
        })
    except Course.DoesNotExist:
        return Response({'success': False, 'message': 'Course not found'}, status=404)


def detect_assessment_type(name):
    """Assessment adından türünü tahmin et"""
    name_lower = name.lower()
    if 'midterm' in name_lower or 'vize' in name_lower:
        return 'midterm'
    elif 'final' in name_lower:
        return 'final'
    elif 'quiz' in name_lower:
        return 'quiz'
    elif 'project' in name_lower or 'proje' in name_lower:
        return 'project'
    elif 'homework' in name_lower or 'ödev' in name_lower:
        return 'homework'
    elif 'attendance' in name_lower or 'devam' in name_lower:
        return 'homework'
    return 'quiz'


@api_view(['POST'])
def import_obs_excel(request):
    """
    OBS Excel formatından ders, öğrenci, assessment ve notları içeri aktar.
    """
    import pandas as pd
    import re
    from io import BytesIO
    
    user = get_user_from_token(request)
    if not user:
        return Response({'success': False, 'message': 'Authorization required'}, status=401)
    
    if 'file' not in request.FILES:
        return Response({'success': False, 'message': 'No file provided'}, status=400)
    
    try:
        excel_file = request.FILES['file']
        df = pd.read_excel(BytesIO(excel_file.read()))
    except Exception as e:
        return Response({'success': False, 'message': f'Excel read error: {str(e)}'}, status=400)
    
    # Course bilgilerini al
    course_id = request.data.get('course_id')
    course_code = request.data.get('course_code', '')
    course_name = request.data.get('course_name', '')
    
    # Sütun isimlerinden course code çıkar (örn: "Öğrenci No_0833AB" -> "0833AB")
    if not course_code:
        for col in df.columns:
            if '_' in col:
                course_code = col.split('_')[-1]
                break
    
    if not course_code:
        return Response({'success': False, 'message': 'Course code not found'}, status=400)
    
    if not course_name:
        course_name = course_code
    
    # Course oluştur veya bul
    created_course = False
    if course_id:
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            return Response({'success': False, 'message': 'Course not found'}, status=404)
    else:
        course, created_course = Course.objects.get_or_create(
            code=course_code,
            defaults={
                'name': course_name,
                'department': 'CSE',
                'instructor': user
            }
        )
    
    # Assessment sütunlarını tespit et: "Assessment Name(%Weight)_CourseCode"
    assessment_pattern = re.compile(r'^(.+?)\(%(\d+)\)_')
    assessment_columns = {}
    
    for col in df.columns:
        match = assessment_pattern.match(col)
        if match:
            assessment_name = match.group(1).strip()
            weight = int(match.group(2))
            assessment_columns[col] = {
                'name': assessment_name,
                'weight': weight
            }
    
    # Assessment'ları oluştur
    created_assessments = 0
    assessments_map = {}
    
    for col, info in assessment_columns.items():
        assessment, created = Assessment.objects.get_or_create(
            course=course,
            name=info['name'],
            defaults={
                'assessment_type': detect_assessment_type(info['name']),
                'total_points': 100  # Default 100 puan üzerinden
            }
        )
        assessments_map[col] = assessment
        if created:
            created_assessments += 1
    
    # Öğrenci sütunlarını bul
    student_no_col = None
    first_name_col = None
    last_name_col = None
    full_name_col = None
    
    for col in df.columns:
        col_lower = col.lower()
        # Öğrenci No sütunu
        if 'öğrenci no' in col_lower or 'student no' in col_lower or 'student_no' in col_lower:
            student_no_col = col
        elif col_lower.startswith('no_'):
            if not student_no_col:
                student_no_col = col
        # Soyadı sütunu (surname/last name) - önce kontrol et
        elif 'soyadı' in col_lower or 'soyadi' in col_lower or 'surname' in col_lower or 'last_name' in col_lower or 'last name' in col_lower:
            last_name_col = col
        # Adı sütunu (first name) - soyadı değilse
        elif ('adı' in col_lower or 'adi' in col_lower or 'first_name' in col_lower or 'first name' in col_lower) and 'soyadı' not in col_lower and 'soyadi' not in col_lower:
            # "Adı Soyadı" birleşik sütun kontrolü
            if 'soyadı' in col_lower or 'soyadi' in col_lower:
                full_name_col = col
            else:
                first_name_col = col
        # Birleşik ad soyad sütunu
        elif 'ad soyad' in col_lower or 'adsoyad' in col_lower or 'full_name' in col_lower or 'fullname' in col_lower:
            full_name_col = col
    
    if not student_no_col:
        # Fallback: İlk sütunda "no" geçiyorsa veya ikinci sütunu kullan
        for col in df.columns:
            if 'no' in col.lower():
                student_no_col = col
                break
        if not student_no_col and len(df.columns) > 1:
            student_no_col = df.columns[0]
    
    if not student_no_col:
        return Response({'success': False, 'message': 'Student ID column not found'}, status=400)
    
    # Öğrenci ve notları işle
    created_students = 0
    updated_students = 0
    created_grades = 0
    
    for index, row in df.iterrows():
        student_no = str(row.get(student_no_col, '')).strip()
        if not student_no or student_no == 'nan':
            continue
        
        first_name = ''
        last_name = ''
        
        # Birleşik ad soyad sütunu varsa, ayır
        if full_name_col:
            full_name = str(row.get(full_name_col, '')).strip()
            if full_name and full_name != 'nan':
                # İsmi boşluktan ayır - ilk parça ad, geri kalan soyad
                name_parts = full_name.split()
                if len(name_parts) >= 2:
                    first_name = name_parts[0]
                    last_name = ' '.join(name_parts[1:])
                elif len(name_parts) == 1:
                    first_name = name_parts[0]
        else:
            # Ayrı sütunlar
            first_name = str(row.get(first_name_col, '')).strip() if first_name_col else ''
            last_name = str(row.get(last_name_col, '')).strip() if last_name_col else ''
        
        if first_name == 'nan':
            first_name = ''
        if last_name == 'nan':
            last_name = ''
        
        # Öğrenci oluştur veya güncelle
        try:
            student = Student.objects.get(student_no=student_no)
            # Mevcut öğrenciyi güncelle - isim bilgisi varsa
            if first_name or last_name:
                if first_name:
                    student.user.first_name = first_name
                if last_name:
                    student.user.last_name = last_name
                student.user.save()
            updated_students += 1
        except Student.DoesNotExist:
            # Yeni kullanıcı ve öğrenci oluştur
            username = f"student_{student_no}"
            user, _ = User.objects.get_or_create(
                username=username,
                defaults={
                    'first_name': first_name or '',
                    'last_name': last_name or '',
                    'email': f"{student_no}@student.edu"
                }
            )
            student = Student.objects.create(
                user=user,
                student_no=student_no,
                department='CSE'
            )
            created_students += 1
        
        # Enrollment oluştur
        Enrollment.objects.get_or_create(student=student, course=course)
        
        # Notları kaydet
        for col, assessment in assessments_map.items():
            try:
                grade_value = row.get(col)
                if pd.notna(grade_value):
                    points = float(grade_value)
                    grade, g_created = Grade.objects.get_or_create(
                        student=student,
                        assessment=assessment,
                        defaults={'points': points}
                    )
                    if not g_created and grade.points != points:
                        grade.points = points
                        grade.save()
                    if g_created:
                        created_grades += 1
            except (ValueError, TypeError):
                pass
    
    return Response({
        'success': True,
        'course_code': course.code,
        'course_created': created_course,
        'students_created': created_students,
        'students_updated': updated_students,
        'assessments_created': created_assessments,
        'grades_created': created_grades
    })


# ============ REPORT GENERATION VIEWS ============

@api_view(['GET'])
def generate_course_report(request, course_id):
    """
    Kurs bazlı rapor oluştur - Excel formatında
    Tüm öğrencilerin PO skorlarını ve assessment notlarını içerir
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response({'error': 'openpyxl kütüphanesi yüklü değil'}, status=500)
    
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({'error': 'Kurs bulunamadı'}, status=404)
    
    # Workbook oluştur
    wb = openpyxl.Workbook()
    
    # ============ SHEET 1: Öğrenci PO Skorları ============
    ws_po = wb.active
    ws_po.title = "PO Scores"
    
    # Stiller
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # PO'ları al
    pos = ProgramOutcome.objects.all().order_by('code')
    
    # Başlıkları yaz
    headers = ['Student No', 'First Name', 'Last Name']
    for po in pos:
        headers.append(po.code)
    headers.append('Average')
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_po.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Kursa kayıtlı öğrencileri al
    enrollments = Enrollment.objects.filter(course=course).select_related('student', 'student__user')
    
    row_idx = 2
    for enrollment in enrollments:
        student = enrollment.student
        
        # Öğrenci bilgileri
        ws_po.cell(row=row_idx, column=1, value=student.student_no).border = thin_border
        ws_po.cell(row=row_idx, column=2, value=student.user.first_name).border = thin_border
        ws_po.cell(row=row_idx, column=3, value=student.user.last_name).border = thin_border
        
        # PO skorlarını hesapla
        po_scores = []
        for col_idx, po in enumerate(pos, 4):
            score = calculate_student_po_score(student, po, course)
            cell = ws_po.cell(row=row_idx, column=col_idx, value=round(score, 2) if score else 0)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if score:
                po_scores.append(score)
        
        # Ortalama
        avg_score = sum(po_scores) / len(po_scores) if po_scores else 0
        avg_cell = ws_po.cell(row=row_idx, column=len(headers), value=round(avg_score, 2))
        avg_cell.border = thin_border
        avg_cell.alignment = Alignment(horizontal="center")
        avg_cell.font = Font(bold=True)
        
        row_idx += 1
    
    # Sütun genişliklerini ayarla
    ws_po.column_dimensions['A'].width = 15
    ws_po.column_dimensions['B'].width = 15
    ws_po.column_dimensions['C'].width = 15
    for i in range(4, len(headers) + 1):
        ws_po.column_dimensions[get_column_letter(i)].width = 10
    
    # ============ SHEET 2: Assessment Notları ============
    ws_grades = wb.create_sheet(title="Assessment Grades")
    
    # Assessmentları al
    assessments = Assessment.objects.filter(course=course).order_by('date', 'name')
    
    # Başlıklar
    grade_headers = ['Student No', 'First Name', 'Last Name']
    for assessment in assessments:
        grade_headers.append(f"{assessment.name} ({assessment.total_points})")
    grade_headers.append('Total Points')
    grade_headers.append('Percentage')
    
    for col_idx, header in enumerate(grade_headers, 1):
        cell = ws_grades.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    row_idx = 2
    for enrollment in enrollments:
        student = enrollment.student
        
        ws_grades.cell(row=row_idx, column=1, value=student.student_no).border = thin_border
        ws_grades.cell(row=row_idx, column=2, value=student.user.first_name).border = thin_border
        ws_grades.cell(row=row_idx, column=3, value=student.user.last_name).border = thin_border
        
        total_points = 0
        total_max = 0
        
        for col_idx, assessment in enumerate(assessments, 4):
            try:
                grade = Grade.objects.get(student=student, assessment=assessment)
                points = grade.points
            except Grade.DoesNotExist:
                points = 0
            
            cell = ws_grades.cell(row=row_idx, column=col_idx, value=points)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            total_points += points
            total_max += assessment.total_points
        
        # Toplam
        total_cell = ws_grades.cell(row=row_idx, column=len(grade_headers) - 1, value=round(total_points, 2))
        total_cell.border = thin_border
        total_cell.font = Font(bold=True)
        
        # Yüzde
        percentage = (total_points / total_max * 100) if total_max > 0 else 0
        pct_cell = ws_grades.cell(row=row_idx, column=len(grade_headers), value=f"{round(percentage, 1)}%")
        pct_cell.border = thin_border
        pct_cell.font = Font(bold=True)
        
        row_idx += 1
    
    # Sütun genişliklerini ayarla
    ws_grades.column_dimensions['A'].width = 15
    ws_grades.column_dimensions['B'].width = 15
    ws_grades.column_dimensions['C'].width = 15
    
    # ============ SHEET 3: LO-PO Mapping ============
    ws_mapping = wb.create_sheet(title="LO-PO Mapping")
    
    los = LearningOutcome.objects.filter(course=course).order_by('code')
    
    # Başlıklar
    mapping_headers = ['LO Code', 'LO Description']
    for po in pos:
        mapping_headers.append(po.code)
    
    for col_idx, header in enumerate(mapping_headers, 1):
        cell = ws_mapping.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    row_idx = 2
    for lo in los:
        ws_mapping.cell(row=row_idx, column=1, value=lo.code).border = thin_border
        ws_mapping.cell(row=row_idx, column=2, value=lo.description[:50]).border = thin_border
        
        for col_idx, po in enumerate(pos, 3):
            try:
                mapping = LoToPoMapping.objects.get(learning_outcome=lo, program_outcome=po)
                weight = mapping.contribution_weight
            except LoToPoMapping.DoesNotExist:
                weight = ''
            
            cell = ws_mapping.cell(row=row_idx, column=col_idx, value=weight if weight else '')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        
        row_idx += 1
    
    ws_mapping.column_dimensions['A'].width = 12
    ws_mapping.column_dimensions['B'].width = 50
    
    # ============ SHEET 4: Özet İstatistikler ============
    ws_summary = wb.create_sheet(title="Summary")
    
    summary_data = [
        ['Course Report Summary', ''],
        ['', ''],
        ['Course Code', course.code],
        ['Course Name', course.name],
        ['Semester', course.semester or 'N/A'],
        ['Department', course.department],
        ['', ''],
        ['Statistics', ''],
        ['Total Students', enrollments.count()],
        ['Total Learning Outcomes', los.count()],
        ['Total Assessments', assessments.count()],
        ['Total LO-PO Mappings', LoToPoMapping.objects.filter(learning_outcome__course=course).count()],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1 or row_idx == 8:
                cell.font = Font(bold=True, size=14)
            if col_idx == 1:
                cell.font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    
    # Excel dosyasını response olarak dön
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{course.code}_report.xlsx"'
    
    return response


@api_view(['GET'])
def generate_student_report(request, student_id):
    """
    Öğrenci bazlı genel rapor - tüm dersler için PO skorları
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response({'error': 'openpyxl kütüphanesi yüklü değil'}, status=500)
    
    course_id = request.query_params.get('course_id')
    
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        return Response({'error': 'Öğrenci bulunamadı'}, status=404)
    
    wb = openpyxl.Workbook()
    
    # Stiller
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if course_id:
        # Belirli bir ders için rapor
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            return Response({'error': 'Kurs bulunamadı'}, status=404)
        
        courses = [course]
        filename = f"{student.student_no}_{course.code}_report.xlsx"
    else:
        # Tüm dersler için rapor
        enrollments = Enrollment.objects.filter(student=student).select_related('course')
        courses = [e.course for e in enrollments]
        filename = f"{student.student_no}_full_report.xlsx"
    
    # ============ SHEET 1: Öğrenci Bilgileri ve Özet ============
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_data = [
        ['Student Report', ''],
        ['', ''],
        ['Student No', student.student_no],
        ['First Name', student.user.first_name],
        ['Last Name', student.user.last_name],
        ['Department', student.department],
        ['', ''],
        ['Enrolled Courses', len(courses)],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            if col_idx == 1:
                cell.font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 30
    
    # ============ SHEET 2: PO Skorları (Tüm Dersler) ============
    ws_po = wb.create_sheet(title="PO Scores")
    
    pos = ProgramOutcome.objects.all().order_by('code')
    
    # Başlıklar
    po_headers = ['Course']
    for po in pos:
        po_headers.append(po.code)
    po_headers.append('Average')
    
    for col_idx, header in enumerate(po_headers, 1):
        cell = ws_po.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    row_idx = 2
    overall_po_scores = {po.id: [] for po in pos}
    
    for course in courses:
        ws_po.cell(row=row_idx, column=1, value=f"{course.code}").border = thin_border
        
        course_scores = []
        for col_idx, po in enumerate(pos, 2):
            score = calculate_student_po_score(student, po, course)
            cell = ws_po.cell(row=row_idx, column=col_idx, value=round(score, 2) if score else 0)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if score:
                course_scores.append(score)
                overall_po_scores[po.id].append(score)
        
        # Ders ortalaması
        avg = sum(course_scores) / len(course_scores) if course_scores else 0
        avg_cell = ws_po.cell(row=row_idx, column=len(po_headers), value=round(avg, 2))
        avg_cell.border = thin_border
        avg_cell.font = Font(bold=True)
        
        row_idx += 1
    
    # Genel ortalama satırı
    if len(courses) > 1:
        ws_po.cell(row=row_idx, column=1, value="OVERALL").border = thin_border
        ws_po.cell(row=row_idx, column=1).font = Font(bold=True)
        
        all_scores = []
        for col_idx, po in enumerate(pos, 2):
            scores = overall_po_scores[po.id]
            avg = sum(scores) / len(scores) if scores else 0
            cell = ws_po.cell(row=row_idx, column=col_idx, value=round(avg, 2))
            cell.border = thin_border
            cell.font = Font(bold=True)
            if avg:
                all_scores.append(avg)
        
        overall_avg = sum(all_scores) / len(all_scores) if all_scores else 0
        ws_po.cell(row=row_idx, column=len(po_headers), value=round(overall_avg, 2)).font = Font(bold=True, color="667eea")
    
    ws_po.column_dimensions['A'].width = 15
    
    # ============ SHEET 3: Assessment Detayları ============
    ws_grades = wb.create_sheet(title="Assessment Details")
    
    grade_headers = ['Course', 'Assessment', 'Type', 'Max Points', 'Score', 'Percentage']
    
    for col_idx, header in enumerate(grade_headers, 1):
        cell = ws_grades.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    row_idx = 2
    for course in courses:
        assessments = Assessment.objects.filter(course=course).order_by('date')
        for assessment in assessments:
            ws_grades.cell(row=row_idx, column=1, value=course.code).border = thin_border
            ws_grades.cell(row=row_idx, column=2, value=assessment.name).border = thin_border
            ws_grades.cell(row=row_idx, column=3, value=assessment.get_assessment_type_display()).border = thin_border
            ws_grades.cell(row=row_idx, column=4, value=assessment.total_points).border = thin_border
            
            try:
                grade = Grade.objects.get(student=student, assessment=assessment)
                score = grade.points
                percentage = grade.percentage
            except Grade.DoesNotExist:
                score = 0
                percentage = 0
            
            ws_grades.cell(row=row_idx, column=5, value=score).border = thin_border
            ws_grades.cell(row=row_idx, column=6, value=f"{round(percentage, 1)}%").border = thin_border
            
            row_idx += 1
    
    ws_grades.column_dimensions['A'].width = 12
    ws_grades.column_dimensions['B'].width = 25
    ws_grades.column_dimensions['C'].width = 12
    ws_grades.column_dimensions['D'].width = 12
    ws_grades.column_dimensions['E'].width = 10
    ws_grades.column_dimensions['F'].width = 12
    
    # Excel dosyasını response olarak dön
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def calculate_student_po_score(student, po, course=None):
    """
    Bir öğrencinin belirli bir PO için skorunu hesapla
    Opsiyonel olarak belirli bir ders için filtreleme yapılabilir
    """
    lo_mappings = LoToPoMapping.objects.filter(program_outcome=po)
    
    if course:
        lo_mappings = lo_mappings.filter(learning_outcome__course=course)
    
    total_weighted_score = 0
    total_weight_sum = 0
    
    for lo_map in lo_mappings:
        lo = lo_map.learning_outcome
        lo_po_weight = lo_map.contribution_weight
        
        assess_mappings = AssessmentToLoMapping.objects.filter(learning_outcome=lo)
        
        for assess_map in assess_mappings:
            assessment = assess_map.assessment
            assess_lo_weight = assess_map.contribution_weight / 100  # Yüzdeden orana çevir
            
            try:
                grade = Grade.objects.get(assessment=assessment, student=student)
                score = grade.percentage
                
                contribution = score * assess_lo_weight * lo_po_weight
                weight_factor = assess_lo_weight * lo_po_weight
                
                total_weighted_score += contribution
                total_weight_sum += weight_factor
            except Grade.DoesNotExist:
                continue
    
    if total_weight_sum > 0:
        return total_weighted_score / total_weight_sum
    return 0
